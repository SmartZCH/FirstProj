
##实现一个用python生成c代码的脚本
'''
1、播放时长以ms为单位，用python脚本生成，C代码中也应该这样子写。
2、输入的参数:
    1、每个语音的命名
    2、语音的序号
3、脚本输出的内容:
    1、地址号 (地址从1开始，下个地址等于上个地址号+播放时长，最后地址号需要四舍五入，并且只保留整数) 
    2、播放时长(播放时长，单位为ms)
    3、生成C语言代码，文件名字保存为voice.h
'''
import wave
import os.path
import xlsxwriter #用来写xlsx文件
import openpyxl #用来读xlsx文件
from openpyxl import load_workbook
import math
 
i = int()
# 音频存放文件夹相对路径
filedir = os.getcwd()
# 获取目录下所有文件
files = os.listdir(filedir)
# 获取目录下所有的WAV文件
wav_files = list()
wav_files_time = list()
for i in files:
    if os.path.splitext(i)[1] == '.WAV':
        wav_files.append(i)
# 文件名按照数字大小排序
wav_files.sort(key = lambda x:int(x[:-4]))
# 获取音频时长
for file in wav_files:
    wav_path = filedir+'/'+file
    with wave.open(wav_path,'rb') as f:
        frames = f.getnframes()
        rate = f.getframerate()
        wav_length = frames / float(rate)
        wav_files_time.append(wav_length)
        #wav_length = round(frames / float(rate), 1)
        print(file,"音频长度：",wav_length,"秒")
# excel文件操作
# 获取excel文件名
excel_name = os.getcwd() + '\\null.xlsx'
# 默认读写的方式打开xlsx文件
wb = load_workbook(excel_name)
# 通过Sheet名获取Sheet对象
ws = wb['Sheet1']
# 将播放时长写入null.xlsx文件中的第C列(单位ms)
all_pile_time= list()
c1 = ws["C1"]
i = 0
for cell in list(ws.columns)[c1.column - 1]:
    if cell.row != 1:
        ws[cell.coordinate] = (int)(wav_files_time[i] * 1000)
        all_pile_time.append((int)(wav_files_time[i] * 1000))
        i = i+1     
print(all_pile_time)        
# 将播放时长向上取整后放入到H列
all_play_time_round = list()
h1 = ws["H1"]
i = 0
for cell in list(ws.columns)[h1.column - 1]:
    if cell.row != 1:
        ws[cell.coordinate] = math.ceil(wav_files_time[i])
        all_play_time_round.append(math.ceil(wav_files_time[i]))
        i = i+1  
print(all_play_time_round)
# 生成序号写入A列序号从0开始
all_num = list()
a1 = ws["a1"]
i = 0
for cell in list(ws.columns)[a1.column - 1]:
    if cell.row != 1:
        ws[cell.coordinate] = i + 1
        all_num.append(i + 1)
        i = i+1  
print(all_num)
# 生成地址号放入到b列
all_addr_num = list()
b1 = ws["b1"]
i = 0
for cell in list(ws.columns)[b1.column - 1]:
    if cell.row != 1:
        if cell.row == 2:
            ws[cell.coordinate] = 1
            all_addr_num.append(1)
        else:    
            ws[cell.coordinate] = all_addr_num[i-1] + all_play_time_round[i-1]
            all_addr_num.append(all_addr_num[i-1] + all_play_time_round[i-1])
        i = i + 1  
# 获取G列的所有数据
all_funciton_name = list()
g1 = ws["G1"]
for cell in list(ws.columns)[g1.column - 1]:
    if cell.row != 1:
        all_funciton_name.append(cell.value)
print(all_funciton_name)
# 获取D列的所有数据
all_comment_name = list()
d1 = ws["D1"]
for cell in list(ws.columns)[d1.column - 1]:
    if cell.row != 1:
        all_comment_name.append(cell.value)
print(all_comment_name)
 
# 除了保存文件外,关于xlsx文件已经完成
# 下面开始生成C代码
 
    #代码部分
with open("txt_copy_to_c.txt","w") as f:
    i = 0
    # 生成第一部分C代码
    while(i < len(all_num)):
        #注释的部分
        f.write("\n") #不多一行换行
        var_comment = "/*  */"
        str_list = list(var_comment)
        str_list.insert(3,all_comment_name[i])
        #extern 部分
        var_extern = "extern const struct voice voice_"
        var_extern += str(all_addr_num[i])
        var_extern += "_"
        var_extern += all_funciton_name[i]
        var_extern += ";"
        f.write("".join(str_list))
        f.write("\n") #不多一行换行
        f.write("".join(var_extern))
        i = i+1
    f.write("\r\n") #换行    
    #生成第二部分C代码
    i = 0
    f.write("const struct voice *voice_group_full_table[VOICE_MAX] = {")
    while(i < len(all_num)):
        f.write("\n") #换行
        f.write("\t") #制表符   
        var_array = "["    
        var_array += all_funciton_name[i].upper()
        var_array += "] = &voice_"
        var_array += str(all_addr_num[i])
        var_array += "_"
        var_array += all_funciton_name[i]
        var_array += ","
        f.write("".join(var_array))
        i = i+1
    f.write("\n") #换行
    f.write("};")
    f.write("\r\n") #换行   
    #生成最后一部分代码
    i = 0
    while(i < len(all_num)):
            #注释部分 
        var_comment = "/*  */"
        str_list = list(var_comment)
        str_list.insert(3,all_comment_name[i]) 
        f.write("".join(str_list))        
            #结构体部分
        f.write("\n") #换行
        var_struct =  "static const struct voice voice_"   
        var_struct += str(all_addr_num[i])
        var_struct += "_"   
        var_struct += all_funciton_name[i]
        var_struct += " = {"
        f.write("".join(var_struct))  
        #地址号
        f.write("\n") 
        f.write("	.play_address = ")
        f.write("".join(str(all_addr_num[i])))
        f.write(",") 
        f.write("\n") 
        #播放时长
        f.write("	.play_time =    ")
        f.write("".join(str(all_pile_time[i])))
        f.write(",") 
        f.write("\n") 
        #音量
        f.write("	.play_volume =  PLAY_VOLUME,")
        f.write("\n") #换行
        #播放进度
        f.write("	.play_schedule = PLAY_SCHEDULE,")
        f.write("\n") #换行
        #播放类型
        f.write("	.play_type = ")
        if (all_funciton_name[i].find("text_",0,len(all_funciton_name[1]))) != -1:
            f.write("PLAY_TYPE_REPEATEDLY")   
        else:
            f.write("PLAY_TYPE_SIGNAL")   
        f.write(",") 
        f.write("\n") #换行
        # 该struct的描述
        f.write("	.desc = ")  
        f.write("".join(str(all_comment_name[i])))  
        f.write(",") 
        f.write("\n") #换行    
        f.write("};")    
        f.write("\n") #换行
        i = i+1
  
#保存文件
wb.save(r"D:\sample.xlsx")
 
#暂停一下
#os.system("pause")
 